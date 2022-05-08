import sys
import openpyxl
import info

def cria_xlsx(arquivo, informacoes):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = " "
    dias_da_semana = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]
    colunas = ['B', 'C', 'D', 'E', 'F']
    count_dia = 0

    for dia_semana in dias_da_semana:
        ws[colunas[count_dia]+'1'] = dias_da_semana[count_dia]
        count_dia += 1

    ws['A2'] = "08:15"

    colunas_materias = 0
    linha_atual = 2
    horario_atual = 0

    for i, horario in enumerate(informacoes.horarios):
        index = informacoes.codigos.index(horario)
        ws[colunas[colunas_materias]+str(linha_atual)] = informacoes.nome_materia[index].split(" ")[0]
        colunas_materias += 1
        if i in informacoes.horarios_tempo:
            ws['A'+str(linha_atual+1)] = informacoes.horarios_bruto[horario_atual]
            linha_atual += 1
            colunas_materias = 0
            horario_atual += 1
    wb.save("test.xlsx")

def printa_horarios(arquivo, informacoes):
    dias_da_semana = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]
    count_dia = 0

    for dia_semana in dias_da_semana:
        count_dia += 1
        print(dia_semana, end=" ")
        if count_dia == 5:
            print("\n")

    for i, horario in enumerate(informacoes.horarios):
        index = informacoes.codigos.index(horario)
        print(informacoes.nome_materia[index].split(" ")[0], end=" ")
        if i in informacoes.horarios_tempo:
            print("\n")

    print("\n")

def extrai_info(arquivo):
    return info.Info(arquivo)

if len(sys.argv) < 3:
    print("Não foram dados argumentos suficientes.")
informacoes = extrai_info(sys.argv[2])
if '-p' in sys.argv[1]:
    printa_horarios(sys.argv[2], informacoes)
if '-e' in sys.argv[1]:
    cria_xlsx(sys.argv[2], informacoes)